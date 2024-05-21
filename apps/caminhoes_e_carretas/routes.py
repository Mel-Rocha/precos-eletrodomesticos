import logging
from io import BytesIO
from typing import List

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from fastapi import APIRouter, Query
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi_pagination import Page, add_pagination, paginate

from apps.caminhoes_e_carretas.models import CaminhoesECarretas
from scraping.caminhoes_e_carretas.backhoe import BackhoeExtract
from apps.caminhoes_e_carretas.schema import CaminhoesECarretasSchema
from automation.caminhoes_e_carretas.collect_url import CaminhoesECarretasAutomation

logging.basicConfig(level=logging.INFO)

load_dotenv()

router = APIRouter()


@router.get("/")
async def get_caminhoes_e_carretas_all() -> Page[CaminhoesECarretasSchema]:
    caminhoes_e_carretas_all = await CaminhoesECarretas.all().values()
    return paginate(caminhoes_e_carretas_all)


add_pagination(router)


@router.get("/extract/backhoe_all/")
async def extract_backhoe_all(urls: List[str] = Query(...)):
    e = BackhoeExtract(urls)
    result = e.extract()
    return {"result": result}


@router.get("/backhoe/excel/", response_class=StreamingResponse)
async def backhoe():
    collect_urls = CaminhoesECarretasAutomation()
    collected_urls, metrics, automation_failure_analysis = collect_urls.backhoe_url_all()

    extraction = BackhoeExtract(collected_urls)
    data, extract_failure_analysis, not_price = extraction.extract()

    logging.info(f"Métricas de desempenho ao obter URLs: {metrics}")
    logging.info(f"Falha ao obter URL, dos seguintes anúncios: {automation_failure_analysis}")
    logging.info(f"Falha na extração: {extract_failure_analysis}")
    logging.info(f"Quantidade de Anúncios sem Preço: {len(not_price)}  URLs Correspondentes: {not_price}")

    df = pd.DataFrame(data)

    df['crawling_date'] = pd.to_datetime(df['crawling_date'])
    df.sort_values(by='crawling_date', ascending=False, inplace=True)

    df.drop_duplicates(subset='url', keep='first', inplace=True)

    df.rename(columns={
        'fabricator': 'Fabricante',
        'model': 'Modelo',
        'year': 'Ano',
        'price': 'Preço',
        'worked_hours': 'Horas',
        'url': 'URL',
        'crawling_date': 'Data da Busca'
    }, inplace=True)

    df['Cód. Modelo'] = ''

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    buffer.seek(0)

    book = openpyxl.load_workbook(buffer)
    sheet = book.active

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, row in enumerate(sheet.iter_rows(min_row=2),
                            start=2):
        fill = fill_blue if i % 2 == 0 else fill_white
        for cell in row:
            cell.fill = fill

    tab = Table(displayName="Table1", ref=sheet.dimensions)

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    sheet.add_table(tab)

    buffer = BytesIO()
    book.save(buffer)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="dataframe.xlsx"'})

